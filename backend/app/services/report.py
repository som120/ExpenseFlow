from __future__ import annotations

import base64
from datetime import UTC, datetime
from io import BytesIO, StringIO

from fpdf import FPDF
from openpyxl.styles import Font, PatternFill
from openpyxl import Workbook

from sqlalchemy.orm import Session

from app.repositories.transaction import TransactionRepository
from app.schemas.report import ExportRead, ReportRead, ReportSection
from app.services.analytics import AnalyticsService
from app.services.summary import SummaryService


class ReportService:
    def __init__(self, db: Session):
        self.summary = SummaryService(db)
        self.analytics = AnalyticsService(db)
        self.transactions = TransactionRepository(db)

    def get_report(self, user_id, report_type: str = "monthly") -> ReportRead:
        summary = self.summary.get_summary(user_id)
        analytics = self.analytics.get_analytics(user_id)
        sections = [
            ReportSection(title="Total Income", value=str(summary.total_income)),
            ReportSection(title="Total Expenses", value=str(summary.total_expenses)),
            ReportSection(title="Net Savings", value=str(summary.net_savings)),
            ReportSection(title="Top Spending Category", value=analytics.top_spending_category or "N/A"),
            ReportSection(title="Highest Expense", value=str(analytics.highest_expense)),
        ]
        return ReportRead(
            report_type=report_type,
            sections=sections,
            generated_at=datetime.now(UTC).isoformat(),
        )

    def export_csv(self, user_id) -> ExportRead:
        report = self.get_report(user_id, report_type="csv")
        transactions = self.transactions.list(user_id)
        buffer = StringIO()
        buffer.write("title,value\n")
        for section in report.sections:
            buffer.write(f"{section.title},{section.value}\n")
        buffer.write("\ntransaction_date,description,type,category,amount,my_share,payment_owner\n")
        for transaction in transactions:
            buffer.write(
                f"{transaction.transaction_date},{transaction.description},{transaction.transaction_type.value},"
                f"{transaction.category.name if transaction.category else 'Others'},{transaction.amount},"
                f"{transaction.my_share},{transaction.payment_owner}\n"
            )
        return ExportRead(filename="expenseflow-report.csv", content=buffer.getvalue(), media_type="text/csv", encoding="utf-8")

    def export_excel(self, user_id) -> ExportRead:
        report = self.get_report(user_id, report_type="excel")
        transactions = self.transactions.list(user_id)
        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Summary"
        summary_sheet.append(["Title", "Value"])
        header_fill = PatternFill(fill_type="solid", fgColor="1F2937")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in summary_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for section in report.sections:
            summary_sheet.append([section.title, section.value])
        summary_sheet.column_dimensions["A"].width = 28
        summary_sheet.column_dimensions["B"].width = 20

        transactions_sheet = workbook.create_sheet("Transactions")
        transactions_sheet.append(["Date", "Description", "Type", "Category", "Amount", "My Share", "Payment Owner"])
        for cell in transactions_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for transaction in transactions:
            transactions_sheet.append(
                [
                    str(transaction.transaction_date),
                    transaction.description,
                    transaction.transaction_type.value,
                    transaction.category.name if transaction.category else "Others",
                    float(transaction.amount),
                    float(transaction.my_share),
                    transaction.payment_owner,
                ]
            )
        for column in ["A", "B", "C", "D", "E", "F", "G"]:
            transactions_sheet.column_dimensions[column].width = 18 if column != "B" else 36

        buffer = BytesIO()
        workbook.save(buffer)
        return ExportRead(
            filename="expenseflow-report.xlsx",
            content=base64.b64encode(buffer.getvalue()).decode("ascii"),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            encoding="base64",
        )

    def export_pdf(self, user_id) -> ExportRead:
        report = self.get_report(user_id, report_type="pdf")
        transactions = self.transactions.list(user_id)
        pdf = FPDF()
        pdf.add_page()
        pdf.set_fill_color(16, 185, 129)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", style="B", size=18)
        pdf.cell(0, 12, "ExpenseFlow Report", new_x="LMARGIN", new_y="NEXT", fill=True)
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("Helvetica", size=11)
        pdf.cell(0, 8, f"Generated at: {report.generated_at}", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(3)
        for section in report.sections:
            pdf.set_font("Helvetica", style="B", size=11)
            pdf.cell(55, 8, f"{section.title}:", border=0)
            pdf.set_font("Helvetica", size=11)
            pdf.cell(0, 8, section.value, new_x="LMARGIN", new_y="NEXT")

        pdf.ln(6)
        pdf.set_font("Helvetica", style="B", size=13)
        pdf.cell(0, 8, "Transactions", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", style="B", size=9)
        pdf.set_fill_color(229, 231, 235)
        headers = [("Date", 24), ("Description", 64), ("Type", 24), ("Category", 28), ("Amount", 22), ("My Share", 24)]
        for label, width in headers:
            pdf.cell(width, 8, label, border=1, fill=True)
        pdf.ln()
        pdf.set_font("Helvetica", size=8)
        for transaction in transactions[:40]:
            pdf.cell(24, 8, str(transaction.transaction_date), border=1)
            pdf.cell(64, 8, transaction.description[:34], border=1)
            pdf.cell(24, 8, transaction.transaction_type.value, border=1)
            pdf.cell(28, 8, (transaction.category.name if transaction.category else "Others")[:14], border=1)
            pdf.cell(22, 8, str(transaction.amount), border=1)
            pdf.cell(24, 8, str(transaction.my_share), border=1)
            pdf.ln()

        content = bytes(pdf.output())
        return ExportRead(
            filename="expenseflow-report.pdf",
            content=base64.b64encode(content).decode("ascii"),
            media_type="application/pdf",
            encoding="base64",
        )
